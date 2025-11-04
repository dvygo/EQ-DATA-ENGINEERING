import os
import csv
from datetime import datetime
import re
try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not installed. Installing...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    from openpyxl import load_workbook

def create_csv_folder(symbol):
    """Create CSV folder for the symbol"""
    base_dir = os.path.dirname(__file__)
    csv_dir = os.path.join(base_dir, 'DATA', symbol.lower(), 'CSV')
    os.makedirs(csv_dir, exist_ok=True)
    return csv_dir

def extract_date_from_filename(filename):
    """Extract date from filename and convert to readable format"""
    # Extract date part from filename (e.g., "03Jun2020" from "03Jun2020_1145_...")
    date_match = re.match(r'(\d{2}[A-Za-z]{3}\d{4})', filename)
    if date_match:
        date_str = date_match.group(1)
        try:
            # Parse the date
            date_obj = datetime.strptime(date_str, '%d%b%Y')
            return date_obj.strftime('%d%b%Y')  # Return in same format
        except ValueError:
            return date_str
    return filename.split('_')[0]  # Fallback

def extract_fields_from_excel(excel_filepath):
    """Extract ProfitLoss and BasicEPS fields from Excel file using XBRL format"""
    try:
        # Load Excel file using openpyxl
        workbook = load_workbook(excel_filepath, data_only=True)
        
        # Look for 'Intance Data' sheet (common in XBRL converted files)
        data_sheet = None
        sheet_names = workbook.sheetnames
        
        for sheet_name in sheet_names:
            if 'intance' in sheet_name.lower() or 'instance' in sheet_name.lower() or 'data' in sheet_name.lower():
                data_sheet = workbook[sheet_name]
                break
        
        if not data_sheet:
            data_sheet = workbook[sheet_names[0]]  # Use first sheet as fallback
        
        # Initialize values
        profit_loss = None
        basic_eps = None
        
        # Search for the required fields in the standard XBRL format
        # Format: Sr.No. | Element Name | Period | Unit | Decimals | Fact Value
        for row in range(1, data_sheet.max_row + 1):
            element_name = data_sheet.cell(row=row, column=2).value  # Column B: Element Name
            fact_value = data_sheet.cell(row=row, column=6).value    # Column F: Fact Value
            
            if element_name and isinstance(element_name, str) and fact_value is not None:
                # Look for Profit/Loss fields (multiple variations using OR conditions)
                if profit_loss is None and any(pattern in element_name for pattern in [
                    'ProfitLossForThePeriod',           # Banking sector
                    'ProfitLossForPeriod',              # Non-banking sector
                    'ProfitLossFromOrdinaryActivitiesAfterTax',  # Alternative banking
                    'ProfitOrLossAttributableToOwnersOfParent'   # Consolidated statements
                ]):
                    try:
                        profit_loss = str(float(fact_value))
                    except (ValueError, TypeError):
                        pass
                
                # Look for EPS fields (multiple variations using OR conditions)
                if basic_eps is None and any(pattern in element_name for pattern in [
                    'BasicEarningsPerShareAfterExtraordinaryItems',                    # Banking sector
                    'BasicEarningsLossPerShareFromContinuingAndDiscontinuedOperations', # Non-banking sector
                    'BasicEarningsPerShareBeforeExtraordinaryItems',                   # Alternative banking
                    'BasicEarningsLossPerShareFromContinuingOperations',               # Continuing operations
                    'DilutedEarningsPerShareAfterExtraordinaryItems',                  # Diluted EPS fallback
                    'BasicEarningsPerShare'                                            # Generic EPS
                ]):
                    try:
                        basic_eps = str(float(fact_value))
                    except (ValueError, TypeError):
                        pass
                
                # If both values found, break early
                if profit_loss is not None and basic_eps is not None:
                    break
        
        workbook.close()
        return profit_loss, basic_eps
        
    except Exception as e:
        print(f"[ERROR] Failed to process {os.path.basename(excel_filepath)}: {e}")
        return None, None

def calculate_number_of_shares(profit_loss, basic_eps):
    """Calculate number of shares outstanding"""
    try:
        if profit_loss and basic_eps:
            profit_val = float(profit_loss)
            eps_val = float(basic_eps)
            if eps_val != 0:
                return int(profit_val / eps_val)
        return 0
    except (ValueError, ZeroDivisionError):
        return 0

def extract_all_excel_files(symbol):
    """Extract data from all Excel files for a symbol"""
    base_dir = os.path.dirname(__file__)
    xlsx_dir = os.path.join(base_dir, 'DATA', symbol.lower(), 'XLSX')
    
    if not os.path.exists(xlsx_dir):
        print(f"[ERROR] XLSX directory not found for {symbol}: {xlsx_dir}")
        return []
    
    excel_files = [f for f in os.listdir(xlsx_dir) if f.endswith('.xlsx') and not f.startswith('~$')]
    
    if not excel_files:
        print(f"[ERROR] No Excel files found for {symbol}")
        return []
    
    print(f"Found {len(excel_files)} Excel files to process for {symbol.upper()}")
    
    extracted_data = []
    successful_count = 0
    failed_count = 0
    
    for excel_file in excel_files:
        excel_path = os.path.join(xlsx_dir, excel_file)
        print(f"Processing: {excel_file}")
        
        # Extract date from filename
        filing_date = extract_date_from_filename(excel_file)
        
        # Extract financial fields
        profit_loss, basic_eps = extract_fields_from_excel(excel_path)
        
        if profit_loss is not None and basic_eps is not None:
            # Calculate number of shares
            num_shares = calculate_number_of_shares(profit_loss, basic_eps)
            
            # Check for duplicate dates
            existing_dates = [item['filing_date'] for item in extracted_data]
            if filing_date in existing_dates:
                print(f"[SKIP] Duplicate date found: Date={filing_date} (keeping existing data)")
                continue
            
            extracted_data.append({
                'filing_date': filing_date,
                'profit_loss': profit_loss,
                'basic_eps': basic_eps,
                'num_shares': num_shares
            })
            
            print(f"[OK] Extracted: Date={filing_date}, ProfitLoss={profit_loss}, EPS={basic_eps}, Shares={num_shares}")
            successful_count += 1
        else:
            print(f"[FAIL] No data found in {excel_file}")
            failed_count += 1
    
    print(f"\nExtraction Summary for {symbol.upper()}:")
    print(f"[OK] Successfully processed: {successful_count} files")
    print(f"[FAIL] Failed to extract: {failed_count} files")
    
    return extracted_data

def save_to_csv(symbol, data):
    """Save extracted data to CSV file"""
    csv_dir = create_csv_folder(symbol)
    csv_file = os.path.join(csv_dir, f"{symbol.lower()}.csv")
    
    # Sort data by date in reverse chronological order
    try:
        data.sort(key=lambda x: datetime.strptime(x['filing_date'], '%d%b%Y'), reverse=True)
    except ValueError:
        # If date parsing fails, sort by string
        data.sort(key=lambda x: x['filing_date'], reverse=True)
    
    with open(csv_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        
        # Write header
        writer.writerow(['FilingDate', 'ProfitLossForThePeriod', 'BasicEarningsPerShareAfterExtraordinaryItems', 'NumberOfSharesOutstanding'])
        
        # Write data
        for item in data:
            writer.writerow([
                item['filing_date'],
                item['profit_loss'],
                item['basic_eps'],
                item['num_shares']
            ])
    
    print(f"[INFO] CSV file saved to: {csv_file}")

def get_available_symbols_with_xlsx():
    """Get list of symbols that have XLSX directories"""
    base_dir = os.path.dirname(__file__)
    data_dir = os.path.join(base_dir, 'DATA')
    
    if not os.path.exists(data_dir):
        print("[ERROR] DATA directory not found")
        return []
    
    symbols = []
    for item in os.listdir(data_dir):
        item_path = os.path.join(data_dir, item)
        if os.path.isdir(item_path):
            xlsx_path = os.path.join(item_path, 'XLSX')
            if os.path.exists(xlsx_path) and os.listdir(xlsx_path):
                symbols.append(item.upper())
    
    return sorted(symbols)

def main():
    """Main function to process all symbols"""
    print("NSE Corporate Filings - Financial Data Extractor")
    print("=" * 60)
    
    # Get available symbols
    symbols = get_available_symbols_with_xlsx()
    
    if not symbols:
        print("[ERROR] No symbols with XLSX files found")
        return
    
    print(f"Found {len(symbols)} symbols with XLSX files: {', '.join(symbols)}")
    print()
    
    # Process each symbol
    for i, symbol in enumerate(symbols, 1):
        print(f"[{i}/{len(symbols)}] Processing {symbol}...")
        print("-" * 50)
        
        # Extract data from Excel files
        extracted_data = extract_all_excel_files(symbol)
        
        # Save to CSV
        save_to_csv(symbol, extracted_data)
        print()
    
    print("Final Extraction Summary:")
    print("=" * 70)
    successful_symbols = []
    failed_symbols = []
    
    for symbol in symbols:
        csv_file = os.path.join(os.path.dirname(__file__), 'DATA', symbol.lower(), 'CSV', f"{symbol.lower()}.csv")
        if os.path.exists(csv_file):
            with open(csv_file, 'r') as f:
                lines = f.readlines()
                if len(lines) > 1:  # More than just header
                    successful_symbols.append(symbol)
                else:
                    failed_symbols.append(symbol)
        else:
            failed_symbols.append(symbol)
    
    print(f"[OK] Successfully processed: {len(successful_symbols)} symbols")
    print(f"[FAIL] Failed to process: {len(failed_symbols)} symbols")
    print(f"[INFO] CSV files organized in DATA/{{symbol}}/CSV/ folders")
    print(f"[INFO] Each CSV contains: FilingDate, ProfitLoss, BasicEPS, NumberOfSharesOutstanding")

if __name__ == "__main__":
    main()
