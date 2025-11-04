import os
import csv
from datetime import datetime
import re
try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not installed. Installing...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    from openpyxl import load_workbook

def create_csv_folder(symbol):
    """Create CSV folder for the symbol"""
    base_dir = os.path.dirname(__file__)
    csv_dir = os.path.join(base_dir, 'DATA', symbol.lower(), 'CSV')
    os.makedirs(csv_dir, exist_ok=True)
    return csv_dir

def extract_date_from_filename(filename):
    """Extract date from filename and convert to readable format"""
    # Extract date part from filename (e.g., "03Jun2020" from "03Jun2020_1145_...")
    date_match = re.match(r'(\d{2}[A-Za-z]{3}\d{4})', filename)
    if date_match:
        date_str = date_match.group(1)
        try:
            # Parse the date
            date_obj = datetime.strptime(date_str, '%d%b%Y')
            return date_obj.strftime('%d%b%Y')  # Return in same format
        except ValueError:
            return date_str
    return filename.split('_')[0]  # Fallback

def extract_fields_from_excel(excel_filepath):
    """Extract ProfitLoss and BasicEPS fields from Excel file"""
    try:
        # Load Excel file using openpyxl
        workbook = load_workbook(excel_filepath, data_only=True)
        
        # Initialize values
        profit_loss = None
        basic_eps = None
        
        # Search in all worksheets
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Search through all cells
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    
                    cell_value = str(cell.value).strip()
                    
                    # Look for ProfitLossForThePeriod
                    if 'ProfitLossForThePeriod' in cell_value or 'Profit (Loss) for the period' in cell_value:
                        # Check adjacent cells for the value
                        row_num = cell.row
                        col_num = cell.column
                        
                        # Check cells to the right
                        for offset in range(1, 5):
                            try:
                                adjacent_cell = worksheet.cell(row=row_num, column=col_num + offset)
                                if adjacent_cell.value is not None:
                                    val = str(adjacent_cell.value).strip()
                                    # Handle negative values and decimal numbers
                                    if re.match(r'^-?\d+\.?\d*$', val.replace(',', '')) or re.match(r'^\(\d+\.?\d*\)$', val.replace(',', '')):
                                        # Convert parentheses format to negative
                                        if val.startswith('(') and val.endswith(')'):
                                            val = '-' + val[1:-1]
                                        profit_loss = val.replace(',', '')  # Remove commas
                                        break
                            except:
                                continue
                    
                    # Look for BasicEarningsPerShareAfterExtraordinaryItems
                    if ('BasicEarningsPerShareAfterExtraordinaryItems' in cell_value or 
                        'Basic earnings per share' in cell_value or
                        'Earnings Per Share' in cell_value):
                        # Check adjacent cells for the value
                        row_num = cell.row
                        col_num = cell.column
                        
                        # Check cells to the right
                        for offset in range(1, 5):
                            try:
                                adjacent_cell = worksheet.cell(row=row_num, column=col_num + offset)
                                if adjacent_cell.value is not None:
                                    val = str(adjacent_cell.value).strip()
                                    # Handle negative values and decimal numbers
                                    if re.match(r'^-?\d+\.?\d*$', val.replace(',', '')) or re.match(r'^\(\d+\.?\d*\)$', val.replace(',', '')):
                                        # Convert parentheses format to negative
                                        if val.startswith('(') and val.endswith(')'):
                                            val = '-' + val[1:-1]
                                        basic_eps = val.replace(',', '')  # Remove commas
                                        break
                            except:
                                continue
                    
                    # Break if both values found
                    if profit_loss and basic_eps:
                        break
                
                if profit_loss and basic_eps:
                    break
            
            if profit_loss and basic_eps:
                break
        
        workbook.close()
        return profit_loss, basic_eps
        
    except Exception as e:
        print(f"[ERROR] Failed to process {os.path.basename(excel_filepath)}: {e}")
        return None, None

def extract_all_excel_files(symbol):
    """Extract fields from all Excel files for a symbol"""
    base_dir = os.path.dirname(__file__)
    xlsx_dir = os.path.join(base_dir, 'DATA', symbol.lower(), 'XLSX')
    
    # Check if XLSX directory exists
    if not os.path.exists(xlsx_dir):
        print(f"[SKIP] XLSX directory not found for {symbol}: {xlsx_dir}")
        return False
    
    # Create CSV directory
    csv_dir = create_csv_folder(symbol)
    
    # Get all Excel files
    excel_files = [f for f in os.listdir(xlsx_dir) if f.endswith('.xlsx') and not f.startswith('~$')]
    
    if not excel_files:
        print(f"[SKIP] No Excel files found for {symbol} in {xlsx_dir}")
        return False
    
    print(f"Found {len(excel_files)} Excel files to process for {symbol}")
    
    # Prepare data for CSV
    extracted_data = []
    processed_count = 0
    failed_count = 0
    
    for excel_file in sorted(excel_files):  # Sort to maintain chronological order
        excel_filepath = os.path.join(xlsx_dir, excel_file)
        
        print(f"Processing: {excel_file}")
        
        # Extract date from filename
        filing_date = extract_date_from_filename(excel_file)
        
        # Extract fields
        profit_loss, basic_eps = extract_fields_from_excel(excel_filepath)
        
        if profit_loss is not None or basic_eps is not None:
            # Check if we already have data for this date
            existing_entry = None
            for i, entry in enumerate(extracted_data):
                if entry[0] == filing_date:
                    existing_entry = i
                    break
            
            # Calculate number of shares outstanding (ProfitLoss / BasicEPS)
            number_of_shares = 'N/A'
            if profit_loss and basic_eps and profit_loss != 'N/A' and basic_eps != 'N/A':
                try:
                    profit_loss_num = float(profit_loss)
                    basic_eps_num = float(basic_eps)
                    if basic_eps_num != 0:
                        number_of_shares = int(abs(profit_loss_num / basic_eps_num))  # Use absolute value for shares
                except (ValueError, ZeroDivisionError):
                    number_of_shares = 'N/A'
            
            new_entry = [
                filing_date,
                profit_loss if profit_loss else 'N/A',
                basic_eps if basic_eps else 'N/A',
                number_of_shares
            ]
            
            if existing_entry is not None:
                # Update existing entry if new data is more complete
                existing = extracted_data[existing_entry]
                if (profit_loss and existing[1] == 'N/A') or (basic_eps and existing[2] == 'N/A'):
                    extracted_data[existing_entry] = new_entry
                    print(f"[UPDATE] Updated existing entry: Date={filing_date}, ProfitLoss={profit_loss}, EPS={basic_eps}, Shares={number_of_shares}")
                else:
                    print(f"[SKIP] Duplicate date found: Date={filing_date} (keeping existing data)")
            else:
                extracted_data.append(new_entry)
                processed_count += 1
                print(f"[OK] Extracted: Date={filing_date}, ProfitLoss={profit_loss}, EPS={basic_eps}, Shares={number_of_shares}")
        else:
            failed_count += 1
            print(f"[FAIL] No data found in {excel_file}")
    
    # Sort data by date in reverse chronological order (newest first)
    def parse_date_for_sorting(date_str):
        """Parse date string for sorting purposes"""
        try:
            return datetime.strptime(date_str, '%d%b%Y')
        except ValueError:
            # Fallback for any date parsing issues
            return datetime.min
    
    extracted_data.sort(key=lambda x: parse_date_for_sorting(x[0]), reverse=True)
    
    # Save to CSV
    csv_filename = f'{symbol.lower()}.csv'
    csv_filepath = os.path.join(csv_dir, csv_filename)
    
    with open(csv_filepath, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        
        # Write header
        writer.writerow(['FilingDate', 'ProfitLossForThePeriod', 'BasicEarningsPerShareAfterExtraordinaryItems', 'NumberOfSharesOutstanding'])
        
        # Write data (now sorted in reverse chronological order)
        writer.writerows(extracted_data)
    
    # Print summary
    print(f"\nExtraction Summary for {symbol}:")
    print(f"[OK] Successfully processed: {processed_count} files")
    print(f"[FAIL] Failed to extract: {failed_count} files")
    print(f"[INFO] CSV file saved to: {csv_filepath}")
    
    return processed_count > 0

def get_available_symbols_with_xlsx():
    """Get list of symbols that have XLSX directories"""
    base_dir = os.path.dirname(__file__)
    data_dir = os.path.join(base_dir, 'DATA')
    
    if not os.path.exists(data_dir):
        print(f"[ERROR] DATA directory not found: {data_dir}")
        return []
    
    symbols = []
    for item in os.listdir(data_dir):
        symbol_dir = os.path.join(data_dir, item)
        xlsx_dir = os.path.join(symbol_dir, 'XLSX')
        
        # Check if it's a directory and has XLSX subdirectory
        if os.path.isdir(symbol_dir) and os.path.exists(xlsx_dir):
            # Check if XLSX directory has any .xlsx files
            xlsx_files = [f for f in os.listdir(xlsx_dir) if f.endswith('.xlsx') and not f.startswith('~$')]
            if xlsx_files:
                symbols.append(item.upper())
    
    return sorted(symbols)

def extract_all_symbols():
    """Extract financial data for all available symbols"""
    print("Starting bulk financial data extraction for all symbols...")
    print("=" * 70)
    
    # Get available symbols with XLSX directories
    symbols = get_available_symbols_with_xlsx()
    if not symbols:
        print("No symbols with XLSX files found. Run converter.py first.")
        return
    
    print(f"Found {len(symbols)} symbols with Excel files to process")
    
    # Process each symbol
    total_successful = 0
    total_failed = 0
    
    for i, symbol in enumerate(symbols, 1):
        print(f"\n[{i}/{len(symbols)}] Processing {symbol}...")
        print("-" * 50)
        
        if extract_all_excel_files(symbol):
            total_successful += 1
        else:
            total_failed += 1
    
    # Print final summary
    print(f"\nFinal Extraction Summary:")
    print(f"=" * 70)
    print(f"[OK] Successfully processed: {total_successful} symbols")
    print(f"[FAIL] Failed to process: {total_failed} symbols")
    print(f"[INFO] CSV files organized in DATA/{{symbol}}/CSV/ folders")
    print(f"[INFO] Each CSV contains: FilingDate, ProfitLoss, BasicEPS, NumberOfSharesOutstanding")

def main():
    """Main function to extract data from Excel files for all symbols"""
    extract_all_symbols()

if __name__ == "__main__":
    main()